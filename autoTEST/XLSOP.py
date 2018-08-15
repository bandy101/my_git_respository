import xlwt
import xlrd
import numpy as np
from xlutils.copy import copy
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.excel import ExcelWriter 
import requests
ip = 'http://plat.dev.ts'
version = 'v1'
prex = ip+'/'+'api/'+version+'/'
js_pwd ={
    'clientId':'098f6bcd4621d373cade4e832627b4f6',
    'userName':'admin',
    'password':'sfe5188'
    }
global token,paramer
def get_token():
    url_login = prex+'login/'
    res = requests.post(url_login,json=js_pwd)
    token = res.json()['content']['token']
    return token
def commit(url,param,METHOD):
    token = get_token()
    if METHOD=='get' or METHOD=='GET':
        res = requests.get(url,params=param,headers={'Authorization':'bearer  '+token})
    if METHOD=='POST' or METHOD=='post':
        res = requests.post(url,json=param,headers={'Content-Type':'application/json','Authorization':'bearer  '+token},timeout=2)
        # print('resurl:',res.url)
    if METHOD=='PUT':
        res = requests.put(url,json=param,headers={'Content-Type':'application/json','Authorization':'bearer  '+token})
        print('url:',res.url)
    r = res.json()
    # return r,r['errmsg']=='OK',res.headers['Date']
    return r,(res.status_code==200 or r['errmsg']=='OK'),res.headers['Date']

def auto(param,METHOD):
    #获取令牌
    token = get_token()
    interface = 'userInfoPageQuery'
    url = prex + interface
    print(url)
    # _ = {'pageNum':1,'pageSize':30}
    # param.update(_)
    
    #-----------------------各种提交------------------------#
    print(METHOD)
    if METHOD=='get' or METHOD=='GET':
        res = requests.get(url,params=param,headers={'Authorization':'bearer  '+token})
    #res = requests.post(url,json=paramer,headers={'Content-Type':'application/json','Authorization':'bearer  '+token})
    # res = requests.put(url,json=paramer,headers={'Content-Type':'application/json','Authorization':'bearer  '+token})
    r = res.json()
    return r,r['errmsg']=='OK',res.headers['Date']
#写录 xls
def start(index=1,names='text',paramer=None):
    # print(paramer)
    xls(paramer,index,names=names)

wbk = xlwt.Workbook()
sheet = wbk.add_sheet('sheet 1')
def xls_add_head(heads,sheet,index=0):
    for i,d in enumerate(heads):
        sheet.write(index,i,d)
    sheet.write(index,len(heads),'结果')
    sheet.write(index,len(heads)+1,'响应时间')

def xls_add_data(datas,sheet,is_ok,date,index=0):
    for i,d in enumerate(datas):
        sheet.write(index,i,d)
    sheet.write(index,len(datas)+1,date)
    if is_ok:   
            sheet.write(index,len(datas),'成功')
    else: 
        sheet.write(index,len(datas),'失败')

def xls(param,index=1,names ='text'):
    re_json,is_ok,date = auto(param,'GET')
    keys = list(param.keys())
    if index <=1:
        wbk = xlwt.Workbook()
        sheet = wbk.add_sheet('sheet 1')
        #添加表头 bt=false
        for i,d in enumerate(keys):
            sheet.write(0,i,d)
            # print('ax',re_json['content']['list'])
            #添加数据
            try:
                # sheet.write(index,i,param[d])
                sheet.write(index,i,re_json['content']['list'][0][d])
            except:
                try:
                    sheet.write(index,i,re_json['content'][d])
                except:pass
        sheet.write(0,len(keys),'结果')
        sheet.write(0,len(keys)+1,'响应时间')
        sheet.write(index,len(keys)+1,date)
        if is_ok:   
            sheet.write(index,len(keys),'成功')
        else: 
            sheet.write(index,len(keys),'失败')
        wbk.save(names+'.xls')
    else:
        workbook = xlrd.open_workbook(names+'.xls')
        workbooknew = copy(workbook)
        ws = workbooknew.get_sheet(0)
        for i,d in enumerate(keys):
            try:
                # ws.write(index,i,param[d])
               ws.write(index,i,re_json['content']['list'][0][d])
            except:
                try:
                    ws.write(index,i,re_json['content'][d])
                except:
                    pass
            
        ws.write(index,len(keys)+1,date)
        if is_ok:
            ws.write(index,len(keys),'成功')
        else: 
            ws.write(index,len(keys),'失败')
        workbooknew.save(names+'.xls')
        return 